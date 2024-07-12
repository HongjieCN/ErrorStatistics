import pandas as pd
import matplotlib.pyplot as plt
from tkinter import filedialog, Tk
import os
from openpyxl import load_workbook
import numpy as np

# 设置Matplotlib字体为微软雅黑
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False

def load_data():
    root = Tk()
    root.withdraw()  # 隐藏Tk窗口
    print("打开文件选择框，请选择Excel文件...")
    file_path = filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel files", "*.xlsx *.xls")])
    if not file_path:
        print("没有选择文件。")
        return None
    directory, file_name = os.path.split(file_path)
    base_name = os.path.splitext(file_name)[0]
    try:
        xl = pd.ExcelFile(file_path)
        data_sheets = [(sheet, pd.read_excel(xl, sheet_name=sheet, header=None)) for sheet in xl.sheet_names]
        return directory, base_name, data_sheets, file_path
    except Exception as e:
        print("加载数据时出错：", e)
        return None


def process_data(df):
    # 确定包含“统计”的列索引，跳过这一列及其右边的所有列
    stats_column = df.iloc[0].eq('统计').idxmax() if '统计' in df.iloc[0].values else len(df.columns)

    # 读取知识点，仅读取至统计列之前
    knowledge_points = df.iloc[1, 1:stats_column].dropna().tolist()

    results = {}
    for index, row in df.iterrows():
        if index > 1:  # 跳过前两行（标题行和知识点行）
            name = row[0]
            errors = row[1:stats_column]  # 读取错误数据直到统计列
            summary = {}
            for kp, error in zip(knowledge_points, errors):
                if pd.notna(error) and error != 0:
                    summary[kp] = summary.get(kp, 0) + 1
            results[name] = summary
    return results, knowledge_points


def plot_histograms(directory, base_name, sheet_name, results):
    sheet_dir = os.path.join(directory, base_name, sheet_name)
    os.makedirs(sheet_dir, exist_ok=True)

    plt.figure(figsize=(15, 8))
    all_knowledge_points = set(kp for r in results.values() for kp in r.keys())
    total_errors = {kp: 0 for kp in all_knowledge_points}

    for errors in results.values():
        for kp, count in errors.items():
            total_errors[kp] += count

    knowledge_points = list(total_errors.keys())
    error_counts = list(total_errors.values())

    bar_width = 0.1
    index = np.arange(len(knowledge_points))

    offset = 0
    for name, errors in results.items():
        counts = [errors.get(kp, 0) for kp in knowledge_points]
        bars = plt.bar(index + offset, counts, bar_width, label=f'{name}', alpha=0.75)
        offset += bar_width
        for bar, count in zip(bars, counts):
            plt.text(bar.get_x() + bar.get_width() / 2, bar.get_height(), f'{count}',
                     ha='center', va='bottom')

    plt.xlabel('知识点')
    plt.ylabel('错误次数')
    plt.title(f'{sheet_name}的错题统计')
    plt.xticks(index + bar_width * len(results) / 2, knowledge_points, rotation=0, ha='center')
    plt.legend()

    plt.tight_layout()
    overall_plot_path = os.path.join(sheet_dir, f"{sheet_name}.png")
    plt.savefig(overall_plot_path)
    plt.close()

    for name, errors in results.items():
        if errors:
            plt.figure(figsize=(10, 6))
            keys = list(errors.keys())
            values = list(errors.values())
            keys_sorted = [x for _, x in sorted(zip(values, keys), reverse=True)]
            values_sorted = sorted(values, reverse=True)
            plt.bar(keys_sorted, values_sorted, color='skyblue', width=0.4)
            plt.xlabel('知识点')
            plt.ylabel('错误次数')
            plt.title(f'{name}的错题直方图')
            plt.xticks(ticks=range(len(keys_sorted)), labels=keys_sorted, rotation=0, ha='center')
            plt.yticks(range(max(values_sorted) + 1))
            plt.tight_layout()
            individual_plot_path = os.path.join(sheet_dir, f"{name}.png")
            plt.savefig(individual_plot_path)
            plt.close()

def update_excel(file_path, sheet_name, results, knowledge_points):
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb[sheet_name]

    # 确定“统计”列位置
    stat_column = None
    for col in range(1, ws.max_column + 2):
        if ws.cell(row=1, column=col).value == "统计":
            stat_column = col
            break
    if not stat_column:
        stat_column = ws.max_column + 2
        ws.cell(row=1, column=stat_column).value = "统计"

    # 初始化知识点列
    kp_start_col = stat_column + 1
    kp_cols = {}
    max_col_used = stat_column

    # 查找或创建知识点列
    for kp in knowledge_points:
        found = False
        for col in range(kp_start_col, ws.max_column + 1):
            if ws.cell(row=2, column=col).value == kp:
                kp_cols[kp] = col
                found = True
                break
        if not found:
            new_col = max_col_used + 1
            ws.cell(row=2, column=new_col).value = kp
            kp_cols[kp] = new_col
            max_col_used = new_col  # 更新使用的最大列数

    # 计算数据应该填写的最大行数
    max_row_used = max((len(results) + 2), ws.max_row)

    # 清除旧数据只在需要更新的单元格
    for col in range(stat_column, max_col_used + 1):
        for row in range(3, max_row_used + 1):
            ws.cell(row=row, column=col).value = None

    # 填充统计数据
    for idx, (name, errors) in enumerate(results.items(), start=3):
        ws.cell(row=idx, column=stat_column).value = name
        for kp, col in kp_cols.items():
            ws.cell(row=idx, column=col).value = errors.get(kp, 0)

    wb.save(file_path)

def main():
    loaded_data = load_data()
    if loaded_data:
        directory, base_name, data_sheets, file_path = loaded_data
        for sheet_name, df in data_sheets:
            results, knowledge_points = process_data(df)
            print(f"生成{sheet_name}的直方图...")
            plot_histograms(directory, base_name, sheet_name, results)
            print(f"更新{sheet_name}的Excel...")
            update_excel(file_path, sheet_name, results, knowledge_points)

if __name__ == "__main__":
    main()
